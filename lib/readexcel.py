#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'Qun Li'

from openpyxl import load_workbook


class ReadExcel():
    """读取excel文件数据"""
    def __init__(self, fileName, SheetName="Sheet1"):
        self.workbook = load_workbook(fileName)
        self.sheet = self.workbook[SheetName]

        # 获取总行数、总列数
        self.nrows = self.sheet.max_row
        self.ncols = self.sheet.max_column

    def read_data(self):
        if self.nrows > 1:
            # 第一行作为 key
            keys = [cell.value for cell in self.sheet[1]]

            listApiData = []
            # 从第二行开始读
            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                api_dict = dict(zip(keys, row))
                listApiData.append(api_dict)

            return listApiData
        else:
            print("表格是空数据!")
            return None
