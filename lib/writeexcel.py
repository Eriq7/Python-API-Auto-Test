#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'Qun Li'

import os
import shutil
from openpyxl import load_workbook
from config import setting


class WriteExcel:
    """写入excel数据"""

    def __init__(self, target_file):
        self.target_file = target_file

        # ✅ 关键修复：确保目标文件的父目录存在
        target_dir = os.path.dirname(self.target_file)
        os.makedirs(target_dir, exist_ok=True)

        # ✅ 如果目标文件不存在：从模板拷贝一份（注意 dst 用 self.target_file）
        if not os.path.exists(self.target_file):
            shutil.copyfile(setting.SOURCE_FILE, self.target_file)

        self.wb = load_workbook(self.target_file)
        self.ws = self.wb.active

    def write_data(self, row_num, value):
        """在指定行写入 Pass/Fail"""
        # 你项目里写第几列按原配置来；这里不改语义
        result_col = 8
        self.ws.cell(row=row_num, column=result_col, value=value)
        self.wb.save(self.target_file)
